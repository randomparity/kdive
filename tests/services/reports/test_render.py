"""CSV and XLSX rendering of a Report (ADR-0208)."""

from __future__ import annotations

import csv
import importlib
import io
from datetime import UTC, datetime

import openpyxl
import pytest

from kdive.domain.errors import CategorizedError, ErrorCategory
from kdive.services.reports import Report, Section
from kdive.services.reports.render import render_csv, render_xlsx

_AS_OF = datetime(2026, 6, 22, 12, 0, tzinfo=UTC)


def _report() -> Report:
    inventory = Section(
        key="inventory",
        columns=("system_id", "vcpus"),
        rows=({"system_id": "s1", "vcpus": 4}, {"system_id": "s2", "vcpus": None}),
        truncated=False,
    )
    leases = Section(
        key="leases",
        columns=("allocation_id",),
        rows=({"allocation_id": "a1"},),
        truncated=True,
    )
    return Report(sections=(inventory, leases), as_of=_AS_OF)


def test_render_csv_one_file_per_section_with_header() -> None:
    out = render_csv(_report())
    assert set(out) == {"inventory", "leases"}
    rows = list(csv.reader(io.StringIO(out["inventory"].decode("utf-8"))))
    assert rows[0] == ["system_id", "vcpus"]
    assert rows[1] == ["s1", "4"]
    assert rows[2] == ["s2", ""]  # None renders as an empty cell


def test_render_csv_marks_truncation() -> None:
    out = render_csv(_report())
    assert b"truncated" in out["leases"].lower()


def test_render_xlsx_sheet_per_section_with_header() -> None:
    workbook = openpyxl.load_workbook(io.BytesIO(render_xlsx(_report())))
    assert workbook.sheetnames == ["inventory", "leases"]
    assert [cell.value for cell in workbook["inventory"][1]] == ["system_id", "vcpus"]


def test_render_xlsx_truncation_note_present() -> None:
    workbook = openpyxl.load_workbook(io.BytesIO(render_xlsx(_report())))
    first_row = [cell.value for cell in workbook["leases"][1]]
    assert first_row[0] is not None and "truncated" in str(first_row[0]).lower()


def test_render_xlsx_missing_openpyxl_reports_optional_extra(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    real_import = importlib.import_module

    def missing_openpyxl(name: str, package: str | None = None) -> object:
        if name == "openpyxl":
            raise ImportError("missing")
        return real_import(name, package)

    monkeypatch.setattr(importlib, "import_module", missing_openpyxl)

    with pytest.raises(CategorizedError) as exc:
        render_xlsx(_report())

    assert exc.value.category is ErrorCategory.MISSING_DEPENDENCY
    assert exc.value.details == {"dependency": "openpyxl", "extra": "report-xlsx"}
